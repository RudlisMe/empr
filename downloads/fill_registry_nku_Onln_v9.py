# -*- coding: utf-8 -*-
"""
Полуавтоматическое заполнение Excel-реестра НКУ.

Версия Onln v9:
  • v9: аргумент registry стал необязательным; без него скрипт сам ищет локальный .xlsx рядом;
  • v9: относительные пути к Excel и JSON-ключу ищутся относительно папки скрипта;
  • добавлена поддержка Google Таблиц по ссылке через service account;
  • онлайн-режим больше не требует openpyxl, если работа идёт только с Google Sheets;
  • онлайн-запись выполняется сразу в Google Sheets;
  • Excel-режим из предыдущих версий сохранён;
  • выбор в меню цифрами;
  • в обычном терминале поддерживаются стрелки ↑/↓ и Enter;
  • Enter повторяет значение из предыдущей строки/предыдущего ввода, где это возможно;
  • автосохранение после каждой полностью добавленной строки;
  • копирование форматирования из предыдущей строки;
  • быстрые команды:
        back / b / назад  — вернуться на шаг назад
        q / quit / exit   — выйти с сохранением уже добавленных строк
        help / h / ?      — показать подсказку
  • цветной и более читаемый вывод в консоли;
  • устойчивое отображение кириллицы на Windows/macOS/Linux.

Как запустить:
    pip install openpyxl
    python fill_registry_nku_Onln_v9.py registry.xlsx
    python fill_registry_nku_Onln_v9.py

По умолчанию результат сохраняется в новый файл registry_filled.xlsx.
Чтобы перезаписать исходный файл с созданием .bak-копии:
    python fill_registry_nku_Onln_v9.py registry.xlsx
    python fill_registry_nku_Onln_v9.py --inplace

Онлайн-режим Google Sheets:
    pip install gspread google-auth
    python fill_registry_nku_Onln_v9.py "https://docs.google.com/spreadsheets/d/.../edit?gid=..." --credentials google_service_account.json
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
from copy import copy
from datetime import date, datetime
from urllib.parse import parse_qs, urlparse
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    load_workbook = None

    def get_column_letter(column: int) -> str:
        """Лёгкая замена openpyxl.utils.get_column_letter для онлайн-режима."""
        if column < 1:
            raise ValueError(f"Некорректный номер колонки: {column}")
        letters = ""
        while column:
            column, remainder = divmod(column - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters

    def column_index_from_string(column_letter: str) -> int:
        """Лёгкая замена openpyxl.utils.column_index_from_string для онлайн-режима."""
        text = str(column_letter).strip().upper()
        if not re.fullmatch(r"[A-Z]+", text):
            raise ValueError(f"Некорректная буква колонки: {column_letter!r}")
        result = 0
        for char in text:
            result = result * 26 + (ord(char) - 64)
        return result


# ---------- Справочники из выдержки ТУ ----------

CONSTRUCTION_G1: List[Tuple[str, str]] = [
    ("Б", "Блок"),
    ("П", "Панель"),
    ("Я", "Ящик"),
    ("Ш", "Шкаф, щит закрытый"),
    ("Щ", "Щит открытый"),
    ("ПР", "Пункт распределительный"),
    ("С", "Пульт (консоль)"),
]

CLASS_G2: List[Tuple[str, str]] = [
    ("1", "НКУ управления, защиты, автоматики, измерения и сигнализации электрических станций и подстанций"),
    ("5", "НКУ управления, в том числе асинхронными электродвигателями трёхфазного тока с короткозамкнутым ротором"),
    ("8", "НКУ ввода и распределения электроэнергии"),
    ("9", "НКУ автоматического управления и регулирования"),
]

GROUP_G3: Dict[str, List[Tuple[str, str]]] = {
    "1": [
        ("1", "НКУ силовые и осветительные с автоматическими выключателями"),
        ("3", "НКУ устройства котла, турбины, технологического оборудования"),
        ("4", "НКУ собственных нужд"),
        ("9", "НКУ вспомогательных хозяйств"),
        ("Х", "НКУ прочие"),
    ],
    "5": [
        ("1", "НКУ с прямым пуском, без реверса и без электрического торможения"),
        ("4", "НКУ с прямым пуском, реверсом и торможением, противовключением"),
        ("9", "НКУ управления несколькими электродвигателями"),
        ("Х", "НКУ прочие"),
    ],
    "8": [
        ("1", "НКУ ввода и распределения переменного тока"),
        ("2", "НКУ ввода и распределения постоянного тока"),
        ("3", "НКУ ввода и распределения переменного тока с автоматическим включением резерва (АВР)"),
        ("Х", "НКУ прочие"),
    ],
    "9": [
        ("1", "НКУ автоматического управления различными механизмами"),
        ("2", "НКУ автоматического регулирования"),
        ("Х", "НКУ прочие"),
    ],
}

CURRENT_G4: Dict[float, str] = {
    0: "00",
    0.1: "10", 0.12: "11", 0.16: "12", 0.2: "13", 0.25: "14", 0.3: "15", 0.4: "16", 0.5: "17", 0.6: "18", 0.8: "19",
    1.0: "20", 1.25: "21", 1.6: "22", 2.0: "23", 2.25: "24", 3.2: "25", 4.0: "26", 5.0: "27", 6.3: "28", 8.0: "29",
    10: "30", 12.5: "31", 16: "32", 20: "33", 22.5: "34", 32: "35", 40: "36", 50: "37", 63: "38", 80: "39",
    100: "40", 125: "41", 160: "42", 200: "43", 225: "44", 320: "45", 400: "46", 500: "47", 630: "48", 800: "49",
    1000: "50", 1250: "51", 1600: "52", 2000: "53", 2500: "54", 3200: "55", 4000: "56", 5000: "57", 6300: "58",
    15: "А0", 30: "А1", 60: "А2", 150: "А3", 300: "А4", 600: "А5", 1500: "А6",
    75: "Б0", 120: "Б1", 750: "Б2", 1200: "Б3",
}

POWER_VOLTAGE_G5: List[Tuple[str, str]] = [
    ("0", "Силовая цепь отсутствует"),
    ("1", "- 110 В"),
    ("2", "- 220 В"),
    ("3", "~ 220 В"),
    ("4", "~ 230 В"),
    ("5", "~ 240 В"),
    ("6", "~ 380 В"),
    ("7", "~ 400 В"),
    ("8", "~ 415 В"),
    ("9", "~ 660 В"),
]

CONTROL_VOLTAGE_G6: List[Tuple[str, str]] = [
    ("0", "Цепь управления отсутствует"),
    ("1", "- 110 В"),
    ("2", "- 220 В"),
    ("3", "~ 110 В"),
    ("4", "~ 220 В"),
    ("5", "~ 230 В"),
    ("6", "~ 240 В"),
    ("7", "~ 380 В"),
    ("8", "~ 400 В"),
    ("9", "~ 415 В"),
    ("А", "- 6 В"),
    ("Б", "- 12 В"),
    ("В", "- 24 В"),
    ("Г", "- 36 В"),
    ("Д", "- 48 В"),
    ("Е", "- 60 В"),
    ("И", "- 125 В"),
    ("К", "Резерв"),
    ("Л", "~ 36 В"),
    ("М", "~ 42 В"),
    ("Н", "~ 127 В"),
]

CLIMATE_OPTIONS: List[Tuple[str, str]] = [
    ("У1", "Умеренный климат, категория размещения 1"),
    ("У2", "Умеренный климат, категория размещения 2"),
    ("У3", "Умеренный климат, категория размещения 3"),
    ("У4", "Умеренный климат, категория размещения 4"),
    ("У5", "Умеренный климат, категория размещения 5"),
    ("УХЛ1", "Умеренный и холодный климат, категория размещения 1"),
    ("УХЛ2", "Умеренный и холодный климат, категория размещения 2"),
    ("УХЛ3", "Умеренный и холодный климат, категория размещения 3"),
    ("УХЛ4", "Умеренный и холодный климат, категория размещения 4"),
    ("УХЛ5", "Умеренный и холодный климат, категория размещения 5"),
    ("ДРУГОЕ", "Ввести вручную"),
]

REQUIRED_COLUMNS = [
    "№", "Номер заказа", "номенклатурный номер", "Тип изделия", "Исполнитель", "Дата внесения записи",
    "Заказчик", "Зав.№", "IP", "Ток, А", "Напряжение, В", "Размер", "Вес",
]


# ---------- Цвета, команды и консоль ----------

class C:
    ENABLED = True
    RESET = "\033[0m"
    BOLD = "\033[1m"
    DIM = "\033[2m"
    RED = "\033[31m"
    GREEN = "\033[32m"
    YELLOW = "\033[33m"
    BLUE = "\033[34m"
    CYAN = "\033[36m"

    @classmethod
    def fmt(cls, text: str, *codes: str) -> str:
        if not cls.ENABLED:
            return text
        return "".join(codes) + text + cls.RESET


class GoBack(Exception):
    pass


class UserQuit(Exception):
    pass


HELP_TEXT = """
Команды можно вводить почти в любом поле:
  Enter              — повторить значение в квадратных скобках, если оно есть
  back / b / назад   — вернуться на шаг назад
  q / quit / exit    — выйти с сохранением уже добавленных строк
  help / h / ?       — показать эту подсказку
"""


def ensure_utf8_console() -> None:
    """Делает вывод/ввод кириллицы устойчивее на Windows, macOS и Linux."""
    if sys.platform.startswith("win"):
        os.system("chcp 65001 > nul 2>&1")

    for stream_name in ("stdin", "stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        if stream is not None and hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8")
            except Exception:
                pass

    if os.environ.get("NO_COLOR"):
        C.ENABLED = False
    elif not sys.stdout.isatty():
        # В перенаправленном выводе ANSI-коды обычно мешают.
        C.ENABLED = False


def print_title(text: str) -> None:
    print("\n" + C.fmt(text, C.BOLD, C.CYAN))


def print_ok(text: str) -> None:
    print(C.fmt(text, C.GREEN))


def print_warn(text: str) -> None:
    print(C.fmt(text, C.YELLOW))


def print_err(text: str) -> None:
    print(C.fmt(text, C.RED))


def show_help() -> None:
    print(C.fmt(HELP_TEXT.strip(), C.DIM))


def normalize_command(raw: str) -> str:
    return raw.strip().lower().replace("ё", "е")


def handle_command(raw: str) -> bool:
    """Возвращает True, если введённая строка была командой."""
    cmd = normalize_command(raw)
    if cmd in {"q", "quit", "exit", "выход", "выйти"}:
        raise UserQuit()
    if cmd in {"back", "b", "назад"}:
        raise GoBack()
    if cmd in {"help", "h", "?", "помощь"}:
        show_help()
        return True
    return False


def input_text(prompt: str, default: Optional[Any] = None, required: bool = True) -> str:
    while True:
        suffix = f" [{default}]" if default not in (None, "") else ""
        raw = input(f"{prompt}{C.fmt(suffix, C.DIM)}: ").strip()

        if raw:
            if handle_command(raw):
                continue
            return raw

        if default not in (None, ""):
            return str(default)

        if not required:
            return ""

        print_warn("Значение не должно быть пустым.")


def input_int(prompt: str, min_value: Optional[int] = None, max_value: Optional[int] = None, default: Optional[int] = None) -> int:
    while True:
        raw = input_text(prompt, default=default, required=(default is None)).replace(",", ".")
        try:
            value = int(float(raw))
        except ValueError:
            print_warn("Введите целое число.")
            continue

        if min_value is not None and value < min_value:
            print_warn(f"Число должно быть не меньше {min_value}.")
            continue
        if max_value is not None and value > max_value:
            print_warn(f"Число должно быть не больше {max_value}.")
            continue
        return value


def input_float(prompt: str, default: Optional[float] = None) -> float:
    while True:
        raw = input_text(prompt, default=default, required=(default is None)).replace(",", ".")
        try:
            return float(raw)
        except ValueError:
            print_warn("Введите число, например 32 или 47.5.")


def input_yes_no(prompt: str, default: bool = False) -> bool:
    hint = "Д/н" if default else "д/Н"
    while True:
        raw = input(f"{prompt} [{hint}]: ").strip()
        if raw:
            if handle_command(raw):
                continue

        value = normalize_command(raw)
        if not value:
            return default
        if value in {"д", "да", "y", "yes"}:
            return True
        if value in {"н", "нет", "n", "no"}:
            return False
        print_warn("Введите 'да' или 'нет'.")


# ---------- Меню со стрелками и цифровой fallback ----------

def _read_key() -> Optional[str]:
    if not sys.stdin.isatty():
        return None

    try:
        if sys.platform.startswith("win"):
            import msvcrt
            ch = msvcrt.getwch()
            if ch in ("\x00", "à"):
                ch2 = msvcrt.getwch()
                if ch2 == "H":
                    return "UP"
                if ch2 == "P":
                    return "DOWN"
                return ch2
            if ch in ("\r", "\n"):
                return "ENTER"
            return ch

        import termios
        import tty
        fd = sys.stdin.fileno()
        old_settings = termios.tcgetattr(fd)
        try:
            tty.setraw(fd)
            ch = sys.stdin.read(1)
            if ch == "\x1b":
                seq = sys.stdin.read(2)
                if seq == "[A":
                    return "UP"
                if seq == "[B":
                    return "DOWN"
                return "ESC"
            if ch in ("\r", "\n"):
                return "ENTER"
            return ch
        finally:
            termios.tcsetattr(fd, termios.TCSADRAIN, old_settings)
    except Exception:
        return None


def _clear_menu_lines(line_count: int) -> None:
    if not sys.stdout.isatty():
        return
    for _ in range(line_count):
        sys.stdout.write("\033[F\033[K")
    sys.stdout.flush()


def index_of_code(options: List[Tuple[str, str]], code: Optional[str], default_index: int = 0) -> int:
    if code is not None:
        for i, (c, _) in enumerate(options):
            if c.upper() == str(code).upper():
                return i
    return max(0, min(default_index, len(options) - 1))


def choose_from_menu_arrows(title: str, options: List[Tuple[str, str]], selected: int) -> Optional[str]:
    if not sys.stdin.isatty() or not sys.stdout.isatty():
        return None

    first_render = True
    while True:
        lines: List[str] = []
        lines.append("\n" + C.fmt(title, C.BOLD, C.CYAN))
        lines.append(C.fmt("↑/↓ — перемещение, Enter — выбрать, цифра — быстрый выбор, b — назад, q — выход", C.DIM))
        for idx, (code, description) in enumerate(options, start=1):
            pointer = C.fmt("➤", C.GREEN, C.BOLD) if idx - 1 == selected else " "
            lines.append(f" {pointer} {idx}. {C.fmt(code, C.BOLD):<12} — {description}")

        if not first_render:
            _clear_menu_lines(len(lines))
        print("\n".join(lines), flush=True)
        first_render = False

        key = _read_key()
        if key is None:
            print()
            return None

        if key == "UP":
            selected = (selected - 1) % len(options)
        elif key == "DOWN":
            selected = (selected + 1) % len(options)
        elif key == "ENTER":
            print()
            return options[selected][0]
        elif key and key.isdigit():
            number = int(key)
            if 1 <= number <= min(9, len(options)):
                print()
                return options[number - 1][0]
        elif key in {"b", "B", "и", "И"}:
            print()
            raise GoBack()
        elif key in {"q", "Q", "й", "Й", "ESC"}:
            print()
            raise UserQuit()


def choose_from_menu(
    title: str,
    options: List[Tuple[str, str]],
    allow_manual_code: bool = False,
    default_code: Optional[str] = None,
    default_index: int = 0,
) -> str:
    selected = index_of_code(options, default_code, default_index)

    selected_by_arrows = choose_from_menu_arrows(title, options, selected)
    if selected_by_arrows is not None:
        return selected_by_arrows

    print_title(title)
    for idx, (code, description) in enumerate(options, start=1):
        default_mark = C.fmt("  ← Enter", C.DIM) if idx - 1 == selected else ""
        print(f"  {idx}. {C.fmt(code, C.BOLD):<12} — {description}{default_mark}")

    prompt = "Выберите номер пункта"
    if allow_manual_code:
        prompt += " или код"
    prompt += ""

    while True:
        raw = input(f"{prompt} [{selected + 1}]: ").strip().upper().replace("X", "Х")
        if raw:
            if handle_command(raw):
                continue
        if not raw:
            return options[selected][0]

        if raw.isdigit():
            number = int(raw)
            if 1 <= number <= len(options):
                return options[number - 1][0]

        if allow_manual_code:
            valid_codes = {code.upper() for code, _ in options}
            if raw in valid_codes:
                return raw

        print_warn("Некорректный выбор. Введите цифру из меню.")


# ---------- Логика НКУ ----------

def current_index_by_nearest_up(current_value: float) -> Tuple[str, float]:
    sorted_currents = sorted(CURRENT_G4.keys())
    for nominal in sorted_currents:
        if current_value <= nominal:
            return CURRENT_G4[nominal], nominal
    raise ValueError(f"Ток {current_value} А больше максимального значения таблицы Г.4 ({sorted_currents[-1]} А).")


def voltage_value_from_code(code: str, options: List[Tuple[str, str]]) -> int:
    text = dict(options).get(code, "")
    digits = "".join(ch for ch in text if ch.isdigit())
    return int(digits) if digits else 0


def code_from_voltage_value(value: Any, options: List[Tuple[str, str]], fallback: str) -> str:
    try:
        voltage = int(float(str(value).replace(",", ".")))
    except Exception:
        return fallback

    for code, text in options:
        digits = "".join(ch for ch in text if ch.isdigit())
        if digits and int(digits) == voltage:
            return code
    return fallback


def build_product_type(defaults: Dict[str, Any], state: Optional[Dict[str, Any]] = None) -> Tuple[str, float, int, Dict[str, Any]]:
    state = dict(state or {})
    print_title("Формирование колонки 'Тип изделия'")

    steps = [
        "product_name",
        "construction",
        "class",
        "group",
        "serial",
        "current",
        "power_voltage",
        "control_voltage",
        "climate",
    ]
    pos = 0

    while pos < len(steps):
        step = steps[pos]
        try:
            if step == "product_name":
                default = state.get(step) or defaults.get(step) or "ЩЭ-3.1"
                state[step] = input_text("Введите начальное название изделия", default=default)

            elif step == "construction":
                default = state.get(step) or defaults.get(step) or "Ш"
                state[step] = choose_from_menu(
                    "Таблица Г.1 — Конструктивное исполнение НКУ",
                    CONSTRUCTION_G1,
                    default_code=default,
                    default_index=3,
                )

            elif step == "class":
                default = state.get(step) or defaults.get(step) or "8"
                state[step] = choose_from_menu(
                    "Таблица Г.2 — Характеристика класса НКУ",
                    CLASS_G2,
                    default_code=default,
                    default_index=2,
                )

            elif step == "group":
                cls = state["class"]
                default = state.get(step) or defaults.get(step) or "1"
                state[step] = choose_from_menu(
                    f"Таблица Г.3 — Группа НКУ для класса {cls}",
                    GROUP_G3[cls],
                    allow_manual_code=True,
                    default_code=default,
                    default_index=0,
                )

            elif step == "serial":
                default = state.get(step) or defaults.get(step) or 1
                state[step] = input_int("Введите порядковый номер в пределах серии от 1 до 99", 1, 99, default=int(default))

            elif step == "current":
                # v6: неверный ток больше таблицы Г.4 больше не валит скрипт через traceback.
                # Пользователь получает понятное предупреждение и повторный запрос значения.
                default = state.get(step) or defaults.get(step)
                while True:
                    state[step] = input_float("Введите номинальный ток силовой цепи, А", default=default)
                    try:
                        current_index, selected_nominal_current = current_index_by_nearest_up(float(state[step]))
                    except ValueError as exc:
                        print_warn(str(exc))
                        print_warn("Введите значение тока из допустимого диапазона таблицы Г.4 или меньше/equal 6300 А.")
                        default = None
                        continue
                    break

                state["current_index"] = current_index
                state["selected_nominal_current"] = selected_nominal_current
                if selected_nominal_current != float(state[step]):
                    print_warn(
                        f"Точного значения {float(state[step]):g} А нет в таблице Г.4. "
                        f"Будет использовано ближайшее большее: {selected_nominal_current:g} А, индекс {current_index}."
                    )
                else:
                    print_ok(f"По таблице Г.4 выбран индекс тока: {current_index}.")

            elif step == "power_voltage":
                default = state.get(step) or defaults.get(step) or "6"
                state[step] = choose_from_menu(
                    "Таблица Г.5 — Модификация НКУ по напряжению силовой цепи",
                    POWER_VOLTAGE_G5,
                    default_code=default,
                    default_index=6,
                )

            elif step == "control_voltage":
                default = state.get(step) or defaults.get(step) or "0"
                state[step] = choose_from_menu(
                    "Таблица Г.6 — Модификация по напряжению цепи управления",
                    CONTROL_VOLTAGE_G6,
                    default_code=default,
                    default_index=0,
                )

            elif step == "climate":
                default = state.get(step) or defaults.get(step) or "У3"
                menu_default = default if default in {c for c, _ in CLIMATE_OPTIONS} else "У3"
                climate_code = choose_from_menu(
                    "Климатическое исполнение и категория размещения",
                    CLIMATE_OPTIONS,
                    default_code=menu_default,
                    default_index=2,
                )
                if climate_code == "ДРУГОЕ":
                    climate_code = input_text("Введите климатическое исполнение вручную", default=default if default != "ДРУГОЕ" else "У3").upper()
                state[step] = climate_code

            pos += 1

        except GoBack:
            if pos == 0:
                raise
            pos -= 1
            print_warn("Возврат на предыдущий шаг.")

    serial_part = f"{int(state['serial']):02d}"
    series_index = f"{state['construction']}{state['class']}{state['group']}{serial_part}"
    current_index = state["current_index"]
    power_voltage_code = state["power_voltage"]
    control_voltage_code = state["control_voltage"]
    climate_code = state["climate"]

    product_type = f"{state['product_name']} НКУ-{series_index}-{current_index}{power_voltage_code}{control_voltage_code} {climate_code}"
    power_voltage_value = voltage_value_from_code(power_voltage_code, POWER_VOLTAGE_G5)

    print_ok(f"Сформированный тип изделия: {C.fmt(product_type, C.BOLD)}")
    return product_type, float(state["current"]), power_voltage_value, state



# ---------- Работа с Google Sheets ----------

GOOGLE_SHEETS_RE = re.compile(r"^https://docs\.google\.com/spreadsheets/d/", re.IGNORECASE)


def is_google_sheet_url(value: str) -> bool:
    return bool(GOOGLE_SHEETS_RE.search(str(value).strip()))


def gid_from_google_sheet_url(url: str) -> Optional[int]:
    """Достаёт gid листа из обычной ссылки Google Sheets."""
    parsed = urlparse(url)
    query_gid = parse_qs(parsed.query).get("gid", [None])[0]
    if query_gid and str(query_gid).isdigit():
        return int(query_gid)
    fragment = parsed.fragment or ""
    if "gid=" in fragment:
        for part in fragment.split("&"):
            if part.startswith("gid=") and part[4:].isdigit():
                return int(part[4:])
    return None


def serialize_sheet_value(value: Any) -> Any:
    """Готовит значения Python к записи в Google Sheets."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return value


class OnlineCell:
    """Минимальная ячейка, совместимая с частью API openpyxl, которую использует скрипт."""

    def __init__(self, sheet: "OnlineWorksheetAdapter", row: int, column: int):
        self.sheet = sheet
        self.row = row
        self.column = column
        self.coordinate = f"{get_column_letter(column)}{row}"
        self.has_style = False
        self._style = None
        self.number_format = None
        self.alignment = None
        self.border = None
        self.fill = None
        self.font = None

    @property
    def value(self) -> Any:
        return self.sheet.get_value(self.row, self.column)

    @value.setter
    def value(self, new_value: Any) -> None:
        self.sheet.set_value(self.row, self.column, new_value)


class OnlineWorksheetAdapter:
    """Адаптер Google Sheets под функции, которые раньше работали с openpyxl worksheet."""

    is_online = True

    def __init__(self, worksheet: Any):
        self.worksheet = worksheet
        self.title = worksheet.title
        self.max_row = int(worksheet.row_count)
        self.max_column = int(worksheet.col_count)
        self.merged_cells = type("MergedCells", (), {"ranges": []})()
        self.row_dimensions: Dict[int, Any] = {}
        self._values: List[List[Any]] = worksheet.get_all_values()
        self._pending: Dict[Tuple[int, int], Any] = {}
        self._refresh_bounds_from_values()

    def _refresh_bounds_from_values(self) -> None:
        if self._values:
            self.max_row = max(self.max_row, len(self._values))
            self.max_column = max(self.max_column, max((len(row) for row in self._values), default=0))

    def get_value(self, row: int, column: int) -> Any:
        if (row, column) in self._pending:
            return self._pending[(row, column)]
        if row <= 0 or column <= 0:
            return None
        if row <= len(self._values) and column <= len(self._values[row - 1]):
            value = self._values[row - 1][column - 1]
            return None if value == "" else value
        return None

    def set_value(self, row: int, column: int, value: Any) -> None:
        value = serialize_sheet_value(value)
        while len(self._values) < row:
            self._values.append([])
        while len(self._values[row - 1]) < column:
            self._values[row - 1].append("")
        self._values[row - 1][column - 1] = value
        self._pending[(row, column)] = value
        self.max_row = max(self.max_row, row)
        self.max_column = max(self.max_column, column)

    def cell(self, row: int, column: int) -> OnlineCell:
        return OnlineCell(self, row, column)

    def flush(self) -> None:
        if not self._pending:
            return
        data = []
        for (row, col), value in sorted(self._pending.items()):
            a1 = f"{get_column_letter(col)}{row}"
            data.append({"range": a1, "values": [[serialize_sheet_value(value)]]})
        self.worksheet.batch_update(data, value_input_option="USER_ENTERED")
        self._pending.clear()


class OnlineWorkbookAdapter:
    """Минимальный workbook с save(), чтобы add_one_record работал и для Excel, и онлайн."""

    is_online = True

    def __init__(self, worksheet_adapter: OnlineWorksheetAdapter):
        self.worksheet_adapter = worksheet_adapter

    def save(self, _: Any = None) -> None:
        self.worksheet_adapter.flush()


def open_google_sheet(url: str, credentials: str, sheet_name: Optional[str] = None, gid: Optional[int] = None) -> Tuple[OnlineWorkbookAdapter, OnlineWorksheetAdapter, str]:
    try:
        import gspread
    except ImportError as exc:
        raise RuntimeError("Для онлайн-режима установите библиотеки: pip install gspread google-auth") from exc

    credentials_path = Path(credentials)
    if not credentials_path.exists():
        raise FileNotFoundError(f"JSON-ключ service account не найден: {credentials_path}")

    gc = gspread.service_account(filename=str(credentials_path))
    spreadsheet = gc.open_by_url(url)
    target_gid = gid if gid is not None else gid_from_google_sheet_url(url)

    if sheet_name:
        worksheet = spreadsheet.worksheet(sheet_name)
    elif target_gid is not None:
        worksheet = spreadsheet.get_worksheet_by_id(target_gid)
        if worksheet is None:
            raise RuntimeError(f"Не найден лист Google Sheets с gid={target_gid}")
    else:
        worksheet = spreadsheet.sheet1

    adapter = OnlineWorksheetAdapter(worksheet)
    return OnlineWorkbookAdapter(adapter), adapter, spreadsheet.title

# ---------- Работа с Excel ----------

def normalize_header(value: Any) -> str:
    """Нормализует заголовок Excel для устойчивого поиска колонок.

    В новых версиях реестра колонки могут смещаться, а в названиях иногда
    появляются переносы строк, неразрывные пробелы или лишние пробелы.
    Поэтому скрипт ищет колонки по тексту заголовка, а не по фиксированным
    буквам столбцов.
    """
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\n", " ").strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", text)


COLUMN_ALIASES: Dict[str, List[str]] = {
    "№": ["№", "номер", "п/п", "n", "no"],
    "Номер заказа": ["Номер заказа", "№ заказа", "заказ"],
    "номенклатурный номер": ["номенклатурный номер", "ном. номер", "номенклатура"],
    "Тип изделия": ["Тип изделия", "наименование", "изделие"],
    "Исполнитель": ["Исполнитель"],
    "Дата внесения записи": ["Дата внесения записи", "дата", "дата записи"],
    "Заказчик": ["Заказчик"],
    "Зав.№": ["Зав.№", "Зав. №", "Заводской №", "Заводской номер", "Зав №"],
    "IP": ["IP", "Степень защиты", "степень защиты ip"],
    "Ток, А": ["Ток, А", "Ток", "I, A", "Iном, А"],
    "Напряжение, В": ["Напряжение, В", "Напряжение", "U, В", "Uном, В"],
    "Размер": ["Размер", "Размеры", "Габарит", "Габариты", "Габаритные размеры"],
    "Вес": ["Вес", "Вес, кг", "Масса", "Масса, кг"],
}


DIMENSION_VALUE_RE = re.compile(r"\d+\s*[xх×*]\s*\d+(?:\s*[xх×*]\s*\d+)?", re.IGNORECASE)
WEIGHT_VALUE_RE = re.compile(r"(?:^|\s)(?:вес\s*)?\d+(?:[,.]\d+)?\s*(?:кг|kg)?(?:\s|$)", re.IGNORECASE)


def merged_parent_text(ws, row: int, col: int) -> str:
    """Возвращает текст верхней левой ячейки объединённого диапазона, если ячейка внутри него."""
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return normalize_header(ws.cell(merged_range.min_row, merged_range.min_col).value)
    return ""


def column_header_context(ws, header_row: int, col: int, radius: int = 3) -> str:
    """Собирает контекст заголовков по колонке, включая родительские объединённые ячейки."""
    parts: List[str] = []
    start = max(1, header_row - radius)
    end = min(ws.max_row, header_row + radius)
    for row in range(start, end + 1):
        text = normalize_header(ws.cell(row=row, column=col).value)
        parent = merged_parent_text(ws, row, col)
        if text:
            parts.append(text)
        if parent and parent != text:
            parts.append(parent)
    return " ".join(dict.fromkeys(parts))


def count_dimension_like_values(ws, header_row: int, col: int, scan_rows: int = 80) -> int:
    count = 0
    for row in range(header_row + 1, min(ws.max_row, header_row + scan_rows) + 1):
        value = ws.cell(row=row, column=col).value
        if value not in (None, "") and DIMENSION_VALUE_RE.search(str(value)):
            count += 1
    return count


def count_weight_like_values(ws, header_row: int, col: int, scan_rows: int = 80) -> int:
    count = 0
    for row in range(header_row + 1, min(ws.max_row, header_row + scan_rows) + 1):
        value = ws.cell(row=row, column=col).value
        if value in (None, ""):
            continue
        text = str(value).strip().lower().replace("ё", "е")
        if "вес" in text or "кг" in text or "kg" in text:
            if re.search(r"\d", text):
                count += 1
    return count


def refine_dimension_columns_by_existing_data(ws, header_row: int, cols: Dict[str, int]) -> None:
    """
    v5: уточняет колонки размера и веса по фактическим данным под заголовками.

    Причина правки: в некоторых версиях реестра есть многострочные/объединённые
    заголовки. Из-за этого Excel может визуально иметь "Шифр заказа" в L,
    а реальный столбец "Размеры" — дальше, например в P. Простое чтение первой
    найденной ячейки "Размер" в таком случае приводит к записи не в тот столбец.
    """
    max_col = ws.max_column

    size_scores: Dict[int, int] = {col: count_dimension_like_values(ws, header_row, col) for col in range(1, max_col + 1)}
    weight_scores: Dict[int, int] = {col: count_weight_like_values(ws, header_row, col) for col in range(1, max_col + 1)}

    # Для размера выбираем колонку, где уже есть значения вида 260x340x120.
    best_size_col = max(size_scores, key=lambda c: size_scores[c]) if size_scores else None
    if best_size_col and size_scores[best_size_col] > 0:
        current_col = cols.get("Размер")
        current_score = size_scores.get(current_col, 0) if current_col else 0
        current_context = column_header_context(ws, header_row, current_col) if current_col else ""
        best_context = column_header_context(ws, header_row, best_size_col)

        # Переключаемся, если текущая колонка похожа на шифр/заказ или если по данным
        # другая колонка явно больше похожа на размеры.
        current_looks_wrong = any(word in current_context for word in ("шифр", "код", "заказ"))
        best_has_size_header = any(word in best_context for word in ("размер", "габарит"))
        if current_col is None or current_looks_wrong or size_scores[best_size_col] > current_score or best_has_size_header:
            cols["Размер"] = best_size_col

    # Для веса выбираем колонку, где уже есть значения вида "Вес 7кг" / "7 кг".
    best_weight_col = max(weight_scores, key=lambda c: weight_scores[c]) if weight_scores else None
    if best_weight_col and weight_scores[best_weight_col] > 0:
        current_col = cols.get("Вес")
        current_score = weight_scores.get(current_col, 0) if current_col else 0
        current_context = column_header_context(ws, header_row, current_col) if current_col else ""
        best_context = column_header_context(ws, header_row, best_weight_col)
        current_looks_wrong = any(word in current_context for word in ("шифр", "код", "заказ", "размер", "габарит"))
        best_has_weight_header = any(word in best_context for word in ("вес", "масса"))
        if current_col is None or current_looks_wrong or weight_scores[best_weight_col] > current_score or best_has_weight_header:
            cols["Вес"] = best_weight_col


def apply_manual_column_override(cols: Dict[str, int], column_name: str, column_letter: Optional[str]) -> None:
    """Позволяет принудительно задать колонку, если конкретный шаблон Excel нестандартный."""
    if not column_letter:
        return
    normalized = column_letter.strip().upper()
    if not re.fullmatch(r"[A-Z]+", normalized):
        raise ValueError(f"Некорректная буква колонки для '{column_name}': {column_letter!r}")
    cols[column_name] = column_index_from_string(normalized)


def find_column_by_header_text_anywhere(ws, header_words: Tuple[str, ...], max_scan_rows: int = 20) -> Optional[int]:
    """Ищет колонку по тексту заголовка в верхней части листа, даже если строка заголовков была выбрана неудачно."""
    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        for col in range(1, ws.max_column + 1):
            text = normalize_header(ws.cell(row=row, column=col).value)
            if not text:
                continue
            if any(word in text for word in header_words):
                return col
    return None


def prefer_confirmed_dimension_columns(ws, header_row: int, cols: Dict[str, int]) -> None:
    """
    v6: дополнительная защита от записи размера в колонку шифра/заказа.

    Логика:
    1) Если текущая колонка размера по контексту похожа на шифр/заказ/проект, считаем её ошибочной.
    2) В этом случае ищем настоящую колонку размера по заголовкам во всей верхней части листа.
    3) Если пользовательская структура имеет размер в P, и в P действительно написано
       'Размер'/'Габариты', будет выбрана именно P.
    """
    current_size_col = cols.get("Размер")
    current_context = column_header_context(ws, header_row, current_size_col) if current_size_col else ""
    current_wrong = any(word in current_context for word in ("шифр", "код", "заказ", "проект"))

    # Приоритетно проверяем P, потому что в текущем рабочем шаблоне пользователя
    # фактическая колонка размера находится там. Но используем P только если её
    # заголовок/контекст действительно похож на размер.
    p_col = column_index_from_string("P")
    if p_col <= ws.max_column:
        p_context = column_header_context(ws, header_row, p_col)
        if any(word in p_context for word in ("размер", "габарит")):
            cols["Размер"] = p_col
            return

    header_size_col = find_column_by_header_text_anywhere(ws, ("размер", "габарит"))
    if header_size_col and (current_size_col is None or current_wrong):
        cols["Размер"] = header_size_col

    current_weight_col = cols.get("Вес")
    current_weight_context = column_header_context(ws, header_row, current_weight_col) if current_weight_col else ""
    weight_wrong = any(word in current_weight_context for word in ("шифр", "код", "заказ", "проект", "размер", "габарит"))
    header_weight_col = find_column_by_header_text_anywhere(ws, ("вес", "масса"))
    if header_weight_col and (current_weight_col is None or weight_wrong):
        cols["Вес"] = header_weight_col


def find_header_row_and_columns(ws) -> Tuple[int, Dict[str, int]]:
    alias_to_column_name: Dict[str, str] = {}
    for canonical_name in REQUIRED_COLUMNS:
        aliases = COLUMN_ALIASES.get(canonical_name, [canonical_name])
        for alias in aliases:
            alias_to_column_name[normalize_header(alias)] = canonical_name

    # v4: выбираем не первую строку, где встретились "№" и "Тип изделия",
    # а лучшую строку заголовков по количеству найденных нужных колонок.
    # Это защищает от служебных строк и сдвигов структуры реестра.
    best_row = 0
    best_found: Dict[str, int] = {}
    for row in range(1, min(ws.max_row, 80) + 1):
        found: Dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            h_norm = normalize_header(ws.cell(row=row, column=col).value)
            if h_norm in alias_to_column_name:
                found[alias_to_column_name[h_norm]] = col

        has_minimum = "№" in found and "Тип изделия" in found
        better = len(found) > len(best_found)
        same_but_has_size = len(found) == len(best_found) and "Размер" in found and "Размер" not in best_found
        if has_minimum and (better or same_but_has_size):
            best_row = row
            best_found = found

    if best_row:
        # Аварийная привязка под текущую структуру: если есть "Напряжение, В" и "Вес",
        # но почему-то не найден "Размер", то размер обычно стоит между ними.
        if "Размер" not in best_found and "Напряжение, В" in best_found and "Вес" in best_found:
            possible_size_col = best_found["Напряжение, В"] + 1
            if possible_size_col < best_found["Вес"]:
                best_found["Размер"] = possible_size_col

        refine_dimension_columns_by_existing_data(ws, best_row, best_found)
        prefer_confirmed_dimension_columns(ws, best_row, best_found)
        return best_row, best_found

    raise RuntimeError("Не нашёл строку заголовков. Проверьте, что в таблице есть колонки '№' и 'Тип изделия'.")


def column_map_text(cols: Dict[str, int]) -> str:
    """Возвращает человекочитаемую карту найденных колонок для проверки структуры."""
    parts = []
    for name in REQUIRED_COLUMNS:
        if name in cols:
            parts.append(f"{name}={get_column_letter(cols[name])}")
    return ", ".join(parts)


def last_row_with_values(ws, max_col: int) -> int:
    last = 0
    for row in range(1, ws.max_row + 1):
        if any(ws.cell(row=row, column=col).value not in (None, "") for col in range(1, max_col + 1)):
            last = row
    return last


def next_number(ws, header_row: int, col: int) -> int:
    numbers = []
    for row in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=row, column=col).value
        if isinstance(v, int):
            numbers.append(v)
        elif isinstance(v, float) and v.is_integer():
            numbers.append(int(v))
        elif isinstance(v, str) and v.strip().isdigit():
            numbers.append(int(v.strip()))
    return (max(numbers) + 1) if numbers else 1


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    if src_row < 1 or src_row == dst_row:
        return

    # В Google Sheets онлайн-режиме v7 копируем только данные, без форматирования.
    # Формат листа обычно уже применён к колонкам/строкам шаблона.
    if getattr(ws, "is_online", False):
        return

    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height

    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)

        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.border:
            dst.border = copy(src.border)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)


def safe_cell(ws, row: int, cols: Dict[str, int], name: str) -> Any:
    if name not in cols or row < 1:
        return None
    return ws.cell(row=row, column=cols[name]).value


def set_cell(ws, row: int, cols: Dict[str, int], name: str, value: Any) -> None:
    if name in cols:
        ws.cell(row=row, column=cols[name]).value = value


def cell_address(ws, row: int, cols: Dict[str, int], name: str) -> str:
    if name not in cols:
        return "не найдена"
    return ws.cell(row=row, column=cols[name]).coordinate


def verify_written_cell(ws, row: int, cols: Dict[str, int], name: str, expected: Any) -> None:
    if name not in cols:
        print_warn(f"Колонка '{name}' не найдена — значение не записано.")
        return
    actual = ws.cell(row=row, column=cols[name]).value
    if expected not in (None, "") and actual in (None, ""):
        print_err(f"ВНИМАНИЕ: '{name}' введено как {expected!r}, но ячейка {cell_address(ws, row, cols, name)} осталась пустой.")
    else:
        print(C.fmt(f"{name}: {cell_address(ws, row, cols, name)} = {actual}", C.DIM))


def parse_product_type(value: Any) -> Dict[str, Any]:
    result: Dict[str, Any] = {}
    if not value:
        return result

    text = str(value).strip()
    if " НКУ-" not in text:
        return result

    product_name, rest = text.split(" НКУ-", 1)
    result["product_name"] = product_name

    # Ожидаем что-то вида: Ш8111-3560 У3
    try:
        index_part, climate = rest.split(" ", 1)
        left, right = index_part.split("-", 1)
    except ValueError:
        return result

    for construction, _ in sorted(CONSTRUCTION_G1, key=lambda item: len(item[0]), reverse=True):
        if left.startswith(construction):
            tail = left[len(construction):]
            if len(tail) >= 4:
                result["construction"] = construction
                result["class"] = tail[0]
                result["group"] = tail[1].replace("X", "Х")
                try:
                    result["serial"] = int(tail[2:4])
                except ValueError:
                    pass
            break

    if len(right) >= 4:
        result["current_index"] = right[:2]
        result["power_voltage"] = right[2]
        result["control_voltage"] = right[3]

    result["climate"] = climate.strip()
    return result


def previous_defaults_from_row(ws, row: int, cols: Dict[str, int]) -> Dict[str, Any]:
    defaults: Dict[str, Any] = {}

    defaults.update(parse_product_type(safe_cell(ws, row, cols, "Тип изделия")))

    aliases = {
        "order_number": "Номер заказа",
        "ip": "IP",
        "size": "Размер",
        "weight": "Вес",
        "current": "Ток, А",
    }
    for key, column in aliases.items():
        value = safe_cell(ws, row, cols, column)
        if value not in (None, ""):
            defaults[key] = value

    voltage = safe_cell(ws, row, cols, "Напряжение, В")
    if voltage not in (None, ""):
        defaults["power_voltage"] = code_from_voltage_value(voltage, POWER_VOLTAGE_G5, defaults.get("power_voltage", "6"))

    # Хорошие стартовые значения, если таблица пустая или предыдущая строка неполная.
    defaults.setdefault("product_name", "ЩЭ-3.1")
    defaults.setdefault("construction", "Ш")
    defaults.setdefault("class", "8")
    defaults.setdefault("group", "1")
    defaults.setdefault("serial", 1)
    defaults.setdefault("power_voltage", "6")
    defaults.setdefault("control_voltage", "0")
    defaults.setdefault("climate", "У3")

    return defaults


def record_wizard(defaults: Dict[str, Any]) -> Dict[str, Any]:
    """Собирает данные строки. В Excel ничего не пишет, пока строка не собрана полностью."""
    state: Dict[str, Any] = {}
    product_state: Optional[Dict[str, Any]] = None
    steps = ["product_type", "order_number", "ip", "size", "weight"]
    pos = 0

    while pos < len(steps):
        step = steps[pos]
        try:
            if step == "product_type":
                product_type, current, voltage, product_state = build_product_type(defaults, product_state)
                state["product_type"] = product_type
                state["current"] = current
                state["voltage"] = voltage

            elif step == "order_number":
                default = state.get(step) or defaults.get(step)
                state[step] = input_text("Введите номер заказа", default=default)

            elif step == "ip":
                default = state.get(step) or defaults.get(step)
                state[step] = input_text("Введите IP", default=default)

            elif step == "size":
                default = state.get(step) or defaults.get(step)
                state[step] = input_text("Введите размер, например 260x340x120", default=default)

            elif step == "weight":
                default = state.get(step) or defaults.get(step)
                state[step] = input_text("Введите вес, например 7 кг", default=default)

            pos += 1

        except GoBack:
            if pos == 0:
                print_warn("Вы уже на первом шаге ввода строки.")
            else:
                pos -= 1
                print_warn("Возврат на предыдущий шаг.")

    return state


def add_one_record(ws, header_row: int, cols: Dict[str, int], last_data_row: int, save_path: Path, wb) -> Tuple[int, Dict[str, Any]]:
    defaults = previous_defaults_from_row(ws, last_data_row, cols)
    data = record_wizard(defaults)

    row = last_data_row + 1
    # Копируем стиль по всей ширине листа, включая новые колонки после веса/размера.
    max_col = max(ws.max_column, max(cols.values()))
    copy_row_style(ws, last_data_row, row, max_col)

    number_value = next_number(ws, header_row, cols["№"]) if "№" in cols else None
    zav_value = next_number(ws, header_row, cols["Зав.№"]) if "Зав.№" in cols else None

    set_cell(ws, row, cols, "№", number_value)
    set_cell(ws, row, cols, "Номер заказа", data["order_number"])
    # "номенклатурный номер" намеренно пропускается.
    set_cell(ws, row, cols, "Тип изделия", data["product_type"])
    set_cell(ws, row, cols, "Зав.№", zav_value)
    set_cell(ws, row, cols, "IP", data["ip"])
    set_cell(ws, row, cols, "Ток, А", data["current"])
    set_cell(ws, row, cols, "Напряжение, В", data["voltage"])
    set_cell(ws, row, cols, "Размер", data["size"])
    set_cell(ws, row, cols, "Вес", data["weight"])

    if "Дата внесения записи" in cols and ws.cell(row=row, column=cols["Дата внесения записи"]).value is None:
        ws.cell(row=row, column=cols["Дата внесения записи"]).value = datetime.now().date()

    wb.save(save_path)
    if getattr(ws, "is_online", False):
        print_ok(f"Строка {row} добавлена и записана онлайн в Google Sheets")
    else:
        print_ok(f"Строка {row} добавлена и автосохранена: {save_path}")
    print(C.fmt("Проверка записи ключевых полей:", C.DIM))
    verify_written_cell(ws, row, cols, "Размер", data.get("size"))
    verify_written_cell(ws, row, cols, "Вес", data.get("weight"))
    if number_value is not None or zav_value is not None:
        print(C.fmt(f"№ {number_value}, Зав.№ {zav_value}", C.BOLD))

    # Эти значения станут дефолтами при следующем вводе в текущей сессии.
    next_defaults = previous_defaults_from_row(ws, row, cols)
    return row, next_defaults


def prepare_save_path(excel_path: Path, inplace: bool, output: Optional[str]) -> Path:
    if inplace:
        backup_path = excel_path.with_suffix(excel_path.suffix + ".bak")
        shutil.copy2(excel_path, backup_path)
        print_ok(f"Создана резервная копия: {backup_path}")
        return excel_path

    return Path(output) if output else excel_path.with_name(excel_path.stem + "_filled" + excel_path.suffix)




def script_dir() -> Path:
    """Папка, где лежит сам Python-скрипт."""
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd()


def resolve_near_script(path_value: Optional[str]) -> Optional[Path]:
    """Возвращает абсолютный путь; относительный путь ищет рядом со скриптом."""
    if not path_value:
        return None
    path = Path(path_value).expanduser()
    if path.is_absolute():
        return path
    cwd_path = Path.cwd() / path
    if cwd_path.exists():
        return cwd_path.resolve()
    return (script_dir() / path).resolve()


def find_default_excel_file() -> Path:
    """Ищет один подходящий .xlsx файл рядом со скриптом для локального режима."""
    base = script_dir()
    candidates = []
    for path in sorted(base.glob("*.xlsx")):
        name_lower = path.name.lower()
        if name_lower.startswith("~$"):
            continue
        if name_lower.endswith(".bak.xlsx") or ".bak" in name_lower:
            continue
        if name_lower.endswith("_filled.xlsx"):
            continue
        candidates.append(path)

    if not candidates:
        raise FileNotFoundError(
            f"Локальный Excel-файл .xlsx не найден рядом со скриптом: {base}"
        )
    if len(candidates) > 1:
        names = "\n".join(f"  - {p.name}" for p in candidates)
        raise RuntimeError(
            "Рядом со скриптом найдено несколько Excel-файлов .xlsx. "
            "Запустите скрипт с явным путём к нужному файлу или оставьте в папке только один файл:\n"
            + names
        )
    return candidates[0]

def main() -> None:
    ensure_utf8_console()

    parser = argparse.ArgumentParser(description="Полуавтоматическое заполнение реестра НКУ: Excel или Google Sheets")
    parser.add_argument("registry", nargs="?", default=None, help="Путь к Excel-файлу .xlsx или ссылка на Google Таблицу. Если не указан — ищется один .xlsx рядом со скриптом.")
    parser.add_argument("--sheet", default=None, help="Имя листа. Для Google Sheets можно использовать имя листа; иначе берётся gid из ссылки.")
    parser.add_argument("--credentials", default="google_service_account.json", help="JSON-ключ service account для Google Sheets")
    parser.add_argument("--gid", type=int, default=None, help="GID листа Google Sheets, если его нужно указать вручную")
    parser.add_argument("--inplace", action="store_true", help="Excel: сохранить изменения в исходный файл, предварительно создав .bak-копию")
    parser.add_argument("--output", default=None, help="Excel: путь для сохранения результата. По умолчанию: *_filled.xlsx")
    parser.add_argument("--size-col", default=None, help="Принудительная колонка для размера, например P")
    parser.add_argument("--weight-col", default=None, help="Принудительная колонка для веса, например Q")
    args = parser.parse_args()

    registry_arg = str(args.registry).strip() if args.registry else ""
    online_mode = is_google_sheet_url(registry_arg) if registry_arg else False
    spreadsheet_title: Optional[str] = None

    if online_mode:
        credentials_path = resolve_near_script(args.credentials)
        wb, ws, spreadsheet_title = open_google_sheet(
            registry_arg,
            str(credentials_path) if credentials_path else args.credentials,
            sheet_name=args.sheet,
            gid=args.gid,
        )
        save_path: Any = "Google Sheets"
    else:
        if load_workbook is None:
            raise RuntimeError(
                "Для Excel-режима установите библиотеку: pip install openpyxl. "
                "Для онлайн-режима openpyxl больше не требуется."
            )
        excel_path = resolve_near_script(registry_arg) if registry_arg else find_default_excel_file()
        if excel_path is None or not excel_path.exists():
            raise FileNotFoundError(f"Файл не найден: {excel_path}")
        wb = load_workbook(excel_path)
        ws = wb[args.sheet] if args.sheet else wb.active
        output_path = str(resolve_near_script(args.output)) if args.output else None
        save_path = prepare_save_path(excel_path, args.inplace, output_path)

    header_row, cols = find_header_row_and_columns(ws)
    apply_manual_column_override(cols, "Размер", args.size_col)
    apply_manual_column_override(cols, "Вес", args.weight_col)
    missing = [name for name in REQUIRED_COLUMNS if name not in cols]
    if missing:
        print_warn("Не найдены некоторые колонки, они будут пропущены: " + ", ".join(missing))

    last_data_row = last_row_with_values(ws, max(cols.values()))
    if last_data_row < header_row:
        last_data_row = header_row

    print_title("Реестр НКУ")
    if online_mode:
        print(f"Открыта Google Таблица: {C.fmt(str(spreadsheet_title or registry_arg), C.BOLD)}")
    else:
        print(f"Открыт файл: {C.fmt(str(excel_path), C.BOLD)}")
    print(f"Лист: {C.fmt(ws.title, C.BOLD)}")
    print(f"Строка заголовков: {header_row}")
    print(f"Найденные колонки: {column_map_text(cols)}")
    print(f"Последняя заполненная строка: {last_data_row}")
    if online_mode:
        print(f"Режим сохранения: {C.fmt('онлайн, сразу в Google Sheets', C.BOLD)}")
    else:
        print(f"Файл сохранения: {C.fmt(str(save_path), C.BOLD)}")
    show_help()

    try:
        while True:
            last_data_row, _ = add_one_record(ws, header_row, cols, last_data_row, save_path, wb)

            if not input_yes_no("Добавить ещё одно наименование?", default=False):
                break

    except UserQuit:
        # Уже добавленные строки сохраняются сразу после ввода.
        # На всякий случай сохраняем текущее состояние книги без неполных строк.
        wb.save(save_path)
        print_warn("Выход по команде пользователя. Уже добавленные строки сохранены.")

    wb.save(save_path)
    if online_mode:
        print_ok("\nГотово. Изменения записаны в Google Sheets.")
    else:
        print_ok(f"\nГотово. Файл сохранён: {save_path}")


if __name__ == "__main__":
    main()
